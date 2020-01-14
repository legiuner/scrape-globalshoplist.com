import requests
import json

import openpyxl
import os
from random import randint
from time import sleep


##11111111111111111111111111111111111111111
def create_excel():

   excel_name = "Output_Scraper.xlsx"
   # retrieve directory of scraper file
   scraperDir = os.path.dirname(os.path.realpath(__file__))
   os.chdir(scraperDir)

   try:

      #create workbook or replace old workbook
      wb = openpyxl.Workbook()
      #create sheet
      wb.create_sheet("Data", 0)
      #remove default sheet
      wb.remove(wb['Sheet'])

      #save workbook
      wb.save(excel_name)

      #select sheet
      ##ws = wb.active
      ws = wb["Data"]

   except PermissionError:
      print("[Errno 13] Permission denied")

   get_data(excel_name, wb, ws)
##11111111111111111111111111111111111111111

##22222222222222222222222222222222222222222
def get_data(excelName, wbook, wsheet):

   getContentPage = True
   halaman = 0
   barisData = 1

   #create header excel
   kolomData = 1
   wsheet.cell(row=barisData, column=kolomData).value = "Page"
   kolomData += 1
   wsheet.cell(row=barisData, column=kolomData).value = "Domain"
   kolomData += 1
   wsheet.cell(row=barisData, column=kolomData).value = "Title"
   kolomData += 1
   wsheet.cell(row=barisData, column=kolomData).value = "Visit/Month"
   wbook.save(excelName)

   while getContentPage:

      random = randint(1,5)
##      print("random = " + str(random))
      sleep(random)

      halaman += 1

      labelHalaman = "Scrape page "
      labelHalaman += str(halaman)
      print(labelHalaman)

      payload = '{"keywords":"","page":'
      payload += str(halaman)
      payload += ',"pageSize":10,"searchFields":["domain","title","description"]}'
           
      headers = {"User-Agent": "Mozilla/5.0 (Windows NT 6.1; rv:72.0) Gecko/20100101 Firefox/72.0",
                  "Accept": "application/json",
                  "Content-Type": "application/json"
                  }

      url = "https://www.globalshoplist.com/api/shop/search"
      page = requests.post(url, data=payload, headers=headers)
##         page = requests.post(url, payload, headers=headers)
##      print(page.status_code)

      if page.status_code == 200:

         doc = page.content
##            print(doc)
         data = json.loads(page.text)
##            print(data)

         for listData in data['list']:
            domain = listData['domain']
##               print(domain)
            try:
               title = listData['title']
               titleStrip = title.strip()
            except:
               titleStrip = ""
               
            visitMonth = listData['visits']['month']      

            #row start from 1 (not 0)
            #column start from 1 (not 0)
            barisData += 1
            kolomData = 1

            wsheet.cell(row=barisData, column=kolomData).value = halaman
            kolomData += 1
            wsheet.cell(row=barisData, column=kolomData).value = domain
            kolomData += 1
            wsheet.cell(row=barisData, column=kolomData).value = titleStrip
            kolomData += 1
            wsheet.cell(row=barisData, column=kolomData).value = visitMonth
            wbook.save(excelName)
      else:
         print("No Data")
         break
         
##22222222222222222222222222222222222222222


##start_start_start_start_start_start_start_start_start_start_
if __name__ == "__main__":
   create_excel()
   print("Scrape end")
##start_start_start_start_start_start_start_start_start_start_

